# cell_inverter.py <filename.xlsx> - inverts the row and column positions of all cells in the spreadsheet



import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import sys

if len(sys.argv) == 2:
	work_book = openpyxl.load_workbook(sys.argv[1])
else:
	print('\ncell_inverter.py <filename.xlsx> - inverts the row and column positions of all cells in the spreadsheet')

sheet = work_book.active
new_sheet = {}

# store every value from every cell in the new_sheet dict with the key being the inverted position
for i in range(1, sheet.max_row+1):
	for cell in sheet[i]:
		new_col = str(get_column_letter(cell.row))
		new_row = str(column_index_from_string(cell.column))
		new_sheet[new_col+new_row] = cell.value

# write the cell values to their inverted positions
for position in new_sheet.keys():
	sheet[position].value = new_sheet[position]

# always save to a new file name
work_book.save('inverted.xlsx')