# multiplication_table.py <max_int> - Creates a .xlsx with a multiplication table who's max integer is the passed aregument


import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string
import sys 


if len(sys.argv) == 2:
	max_int = int(sys.argv[1])
else:
	print('\nmultiplication_table.py <max_int> - Creates a .xlsx with a multiplication table who\'s max integer is the passed aregument')

# create a new workbook obj
wb = openpyxl.Workbook()
# grab the active sheet-- all wb objects are constructed with one
sheet = wb.active
bold_font = Font(bold=True)

# loop through the cells to print out the integers on the top and far left side 
# in bold
for row in sheet['B1':get_column_letter(max_int+1) + str(1)]:
	for cell in row:
		# each int chould be bold for the headers
		cell.font = bold_font
		cell.value = column_index_from_string(cell.column)-1
for row in sheet['A2':'A'+str(max_int+1)]:
	for cell in row:
		cell.font = bold_font
		cell.value = cell.row-1
# sheet.freeze_panes = 'A2'
sheet.freeze_panes = 'B2'

# fill the cells inside of the table with the products of their cords
for row in sheet['B2':get_column_letter(max_int+1) + str(max_int+1)]:
	for cell in row:
		cell.value = (cell.row-1) * (column_index_from_string(cell.column)-1)
		# cell.value = str(cell.row-1) + ' * ' + str(column_index_from_string(cell.column)-1)
wb.save('multiplication_table.xlsx')