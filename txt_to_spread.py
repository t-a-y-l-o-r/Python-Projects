# txt_to_spread.py - Walks the cwd and stores every txt file
# in a single spread sheet. Columns = file, rows = line from file.

import os
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

work_book = openpyxl.Workbook()
sheet = work_book.active


# postiion : line
file_dict = {}
column = 1
row = 1

# walk the cwd
for current_folder, sub_folders, files in os.walk(os.getcwd()):
    for file in files:
        # allows for better column managment
        file_written = False
        row = 1
        position = get_column_letter(column) + str(row)
        # for the sake of simplicity this script only reads txt files
        if '.txt' in file:
            # will auto close the file once the loop is escaped
            with open(current_folder + '/' + file, 'r') as open_file:
                for line in open_file:
                    if line is not '\n':
                        # this should only be true is a file is both a
                        # text file and contains
                        # actual text, not just whitespace
                        file_written = True
                        file_dict.setdefault(position, line)
                        row += 1
                        position = get_column_letter(column) + str(row)
        if file_written:
            column += 1python3 x

# store on the active sheet
for key in file_dict:
    sheet[key] = file_dict[key]


work_book.save('txt_lines.xlsx')
