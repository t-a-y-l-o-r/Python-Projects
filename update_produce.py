# update_produce.py - Corrects the cost in produce sales spreadsheets

import openpyxl

work_book = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# The priduce types and their updated prices 
# making it easier for future changes
PRICE_UPDATES = {
		'Garlic' : 3.07,
		'Celery' : 1.19,
		'Lemon' : 1.27
}

# Loop through the rows and update the prices
for row in range(2, sheet.max_row):	#skip the first row 
	produce_name = sheet.cell(row=row, column=1).value
	if produce_name in PRICE_UPDATES: # god I love python
		# will change the price for the machted name to the price found in our dict
		sheet.cell(row=row, column=2).value = PRICE_UPDATES[produce_name]

# always save to a new file!
work_book.save('updated_produce_sales.xlsx')