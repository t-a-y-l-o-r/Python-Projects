# spread_to_txt.py - Converts all data in a single spread sheet to text files
# columns = file, rows = line in that file

import os

import openpyxl
from openpyxl.utils import get_column_letter


work_book = openpyxl.load_workbook('txt_lines.xlsx')
sheet = work_book.active
max_r = sheet.max_row
max_c = sheet.max_column

os.makedirs('./spread_to_txt', exist_ok=True)


for column_int in range(1, max_c + 1):
	column = get_column_letter(column_int)
	with open('./spread_to_txt/' + str(column) + '.txt', 'w') as file:
		for row in range(1, max_r + 1):
			# No point in writing blank lines to the file
			if sheet[column + str(row)].value:
				file.write(sheet[column + str(row)].value)
				